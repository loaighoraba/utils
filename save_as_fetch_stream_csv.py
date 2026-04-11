#!/usr/bin/env python3
"""
Stream URLs from many Excel files, fetch data with httpx using an asyncio worker pool,
and write results incrementally to a CSV using csv.writer running in a dedicated thread.

This version uses the synchronous csv.writer (faster and CSV-correct) while ensuring
the event loop never blocks by handing CSV writes to a background thread via a
thread-safe queue.Queue.

Usage:
    python fetch_stream_csv_writer.py urls1.xlsx urls2.xlsx ...
"""

import asyncio
import csv
import logging
import threading
import time
from typing import List, Optional
from dataclasses import dataclass

import httpx
from openpyxl import load_workbook
import queue as sync_queue_module


# ================== Configuration ==================
@dataclass(frozen=True)
class Config:
    """Configuration for the fetch-stream-csv pipeline."""

    concurrency: int = 50
    url_queue_maxsize: int = 2000
    result_queue_maxsize: int = 2000
    request_timeout: float = 20.0
    retries: int = 2
    backoff_base: float = 0.5  # seconds
    csv_out: str = "results.csv"
    url_column_index: int = (
        1  # 1-based index of column that contains URL in Excel (adjust)
    )


CONFIG = Config()


# ================== Logging Setup ==================
def setup_logging():
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
    )


setup_logging()


# ================== Excel Utilities ==================
def enqueue_urls_from_excel(
    path: str, loop: asyncio.AbstractEventLoop, url_queue: asyncio.Queue, config: Config
) -> None:
    """
    Reads URLs from an Excel file and schedules them to be put into an asyncio.Queue.
    Runs in a thread.
    """
    logging.info("Reading Excel: %s", path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        url = (
            row[config.url_column_index - 1]
            if len(row) >= config.url_column_index
            else None
        )
        if not url:
            continue
        asyncio.run_coroutine_threadsafe(url_queue.put(str(url).strip()), loop)
    wb.close()
    logging.info("Finished scheduling URLs from: %s", path)


async def produce_from_excels(
    paths: List[str], url_queue: asyncio.Queue, config: Config
) -> None:
    """
    Launches threads to read URLs from multiple Excel files and enqueue them.
    """
    loop = asyncio.get_running_loop()
    for path in paths:
        await asyncio.to_thread(enqueue_urls_from_excel, path, loop, url_queue, config)


# ---------- HTTP fetching with httpx (async) ----------


# ================== Result Dataclass ==================
@dataclass
class FetchResult:
    url: str
    result: str


# ================== HTTP Utilities ==================
async def fetch_with_retries(
    client: httpx.AsyncClient, url: str, config: Config
) -> str:
    """
    Fetches a URL with retries and exponential backoff.
    """
    last_exc: Optional[Exception] = None
    for attempt in range(1, config.retries + 2):
        try:
            r = await client.get(url, timeout=config.request_timeout)
            r.raise_for_status()
            return r.text[:200]  # example: first 200 chars
        except Exception as exc:
            last_exc = exc
            if attempt <= config.retries:
                await asyncio.sleep(config.backoff_base * (2 ** (attempt - 1)))
            else:
                return f"ERROR: {repr(last_exc)}"
    return f"ERROR: {repr(last_exc)}"


# ================== Worker Logic ==================
async def worker(
    worker_id: int,
    url_queue: asyncio.Queue,
    result_sync_queue: sync_queue_module.Queue,
    client: httpx.AsyncClient,
    config: Config,
):
    """
    Async worker that fetches URLs and puts results into a thread-safe queue.
    """
    loop = asyncio.get_running_loop()
    logging.info("Worker %d started", worker_id)
    while True:
        url = await url_queue.get()
        if url is None:
            url_queue.task_done()
            break
        try:
            result = await fetch_with_retries(client, url, config)
            fetch_result = FetchResult(url=url, result=result)
            await loop.run_in_executor(None, result_sync_queue.put, fetch_result)
        except Exception as exc:
            fetch_result = FetchResult(url=url, result=f"ERROR: {exc!r}")
            await loop.run_in_executor(None, result_sync_queue.put, fetch_result)
        finally:
            url_queue.task_done()
    logging.info("Worker %d exiting", worker_id)


# ================== CSV Writer ==================
def csv_writer_thread(
    result_sync_queue: sync_queue_module.Queue, csv_path: str
) -> None:
    """
    Runs in a dedicated thread. Consumes FetchResult objects from result_sync_queue
    and writes them using csv.writer. Stops when it receives a None sentinel.
    """
    logging.info("CSV writer thread starting, writing to %s", csv_path)
    with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(["url", "result"])
        while True:
            item = result_sync_queue.get()  # blocking
            try:
                if item is None:
                    result_sync_queue.task_done()
                    break
                writer.writerow([item.url, str(item.result)])
                result_sync_queue.task_done()
            except Exception:
                result_sync_queue.task_done()
                raise
    logging.info("CSV writer thread finished")


# ================== Orchestration ==================
async def main(
    excel_paths: List[str], out_csv: str = CONFIG.csv_out, config: Config = CONFIG
) -> None:
    """
    Orchestrates the pipeline: reads URLs, fetches them, and writes results to CSV.
    """
    url_queue: asyncio.Queue = asyncio.Queue(maxsize=config.url_queue_maxsize)
    result_sync_queue: sync_queue_module.Queue = sync_queue_module.Queue(
        maxsize=config.result_queue_maxsize
    )

    writer_thread = threading.Thread(
        target=csv_writer_thread, args=(result_sync_queue, out_csv), daemon=False
    )
    writer_thread.start()

    limits = httpx.Limits(
        max_connections=config.concurrency, max_keepalive_connections=config.concurrency
    )
    async with httpx.AsyncClient(limits=limits) as client:
        workers = [
            asyncio.create_task(worker(i, url_queue, result_sync_queue, client, config))
            for i in range(config.concurrency)
        ]

        await produce_from_excels(excel_paths, url_queue, config)

        for _ in range(config.concurrency):
            await url_queue.put(None)

        await url_queue.join()
        await asyncio.gather(*workers)

    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, result_sync_queue.put, None)
    await loop.run_in_executor(None, result_sync_queue.join)
    writer_thread.join()
    logging.info("All done. Results written to %s", out_csv)


# ================== Entrypoint ==================
def cli() -> None:
    import sys

    excel_files = sys.argv[1:]
    if not excel_files:
        print("Usage: python fetch_stream_csv_writer.py file1.xlsx file2.xlsx ...")
        raise SystemExit(1)
    t0 = time.time()
    asyncio.run(main(excel_files, CONFIG.csv_out, CONFIG))
    logging.info("Total elapsed: %.2f seconds", time.time() - t0)


if __name__ == "__main__":
    cli()
